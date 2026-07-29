import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import os


def create_product_template(
    seller_id=None
):

    folder = "uploads/templates"

    os.makedirs(
        folder,
        exist_ok=True
    )


    file_path = os.path.join(
        folder,
        "product_template.xlsx"
    )


    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Products"



    headers = [

        "offer_id",
        "sku",
        "name",
        "description",
        "brand",
        "price",
        "old_price",
        "stock",
        "image"

    ]


    ws.append(headers)



    # Namuna qator

    ws.append([

        "10001",
        "PHONE001",
        "iPhone 15",
        "128GB telefon",
        "Apple",
        12000000,
        13000000,
        10,
        "image_url"

    ])



    # Stil

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="2563eb"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )



    for column in ws.columns:

        max_length = 0

        col = column[0].column


        for cell in column:

            if cell.value:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )


        ws.column_dimensions[
            get_column_letter(col)
        ].width = max_length + 5



    wb.save(
        file_path
    )


    return file_path